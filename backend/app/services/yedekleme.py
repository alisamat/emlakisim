"""
YEDEKLEME SERVİSİ — Excel export, email ile gönderim
Kullanıcı kendi verisinin yedeğinden sorumludur.
"""
import io
import json
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


def excel_export(emlakci):
    """Tüm veriyi Excel formatında export et → bytes döndür."""
    try:
        import openpyxl
    except ImportError:
        return _json_export(emlakci)

    from app.models import Musteri, Mulk, YerGosterme, Not
    from app.models.muhasebe import GelirGider, Cari
    from app.models.planlama import Gorev
    from app.models.fatura import Fatura
    from app.models.lead import Lead
    from app.models.iletisim_gecmisi import IletisimKayit

    wb = openpyxl.Workbook()

    # Müşteriler
    ws = wb.active
    ws.title = 'Musteriler'
    ws.append(['ID', 'Ad Soyad', 'Telefon', 'TC', 'Islem Turu', 'Butce Min', 'Butce Max', 'Sicaklik', 'Notlar', 'Tarih'])
    for m in Musteri.query.filter_by(emlakci_id=emlakci.id).all():
        ws.append([m.id, m.ad_soyad, m.telefon, m.tc_kimlik, m.islem_turu, m.butce_min, m.butce_max, m.sicaklik, m.tercih_notlar, str(m.olusturma or '')])

    # Portföy
    ws2 = wb.create_sheet('Portfoy')
    ws2.append(['ID', 'Baslik', 'Adres', 'Sehir', 'Ilce', 'Tip', 'Islem', 'Fiyat', 'Metrekare', 'Oda', 'Kat', 'Bina Yasi', 'Isitma', 'Esyali', 'Krediye Uygun', 'Sahibi', 'Sahip Tel', 'Detaylar', 'Tarih'])
    for m in Mulk.query.filter_by(emlakci_id=emlakci.id, aktif=True).all():
        d = m.detaylar or {}
        # Mülk sahibi bilgisi (müşteriden)
        sahip_ad, sahip_tel = '', ''
        if m.musteri_id:
            sahip = Musteri.query.get(m.musteri_id)
            if sahip:
                sahip_ad = sahip.ad_soyad or ''
                sahip_tel = sahip.telefon or ''
        ws2.append([m.id, m.baslik, m.adres, m.sehir, m.ilce, m.tip, m.islem_turu, m.fiyat, m.metrekare, m.oda_sayisi,
                    d.get('kat'), d.get('bina_yasi'), d.get('isitma'), d.get('esyali'), d.get('krediye_uygun'),
                    sahip_ad, sahip_tel, json.dumps(d, ensure_ascii=False), str(m.olusturma or '')])

    # Gelir/Gider
    ws3 = wb.create_sheet('GelirGider')
    ws3.append(['ID', 'Tip', 'Kategori', 'Tutar', 'Aciklama', 'Tarih'])
    for k in GelirGider.query.filter_by(emlakci_id=emlakci.id).all():
        ws3.append([k.id, k.tip, k.kategori, k.tutar, k.aciklama, str(k.tarih or '')])

    # Görevler
    ws4 = wb.create_sheet('Gorevler')
    ws4.append(['ID', 'Baslik', 'Tip', 'Oncelik', 'Durum', 'Baslangic', 'Aciklama'])
    for g in Gorev.query.filter_by(emlakci_id=emlakci.id).all():
        ws4.append([g.id, g.baslik, g.tip, g.oncelik, g.durum, str(g.baslangic or ''), g.aciklama])

    # Notlar
    ws5 = wb.create_sheet('Notlar')
    ws5.append(['ID', 'Icerik', 'Etiket', 'Tarih'])
    for n in Not.query.filter_by(emlakci_id=emlakci.id).all():
        ws5.append([n.id, n.icerik, n.etiket, str(n.olusturma or '')])

    # Yer Göstermeler
    ws6 = wb.create_sheet('YerGostermeler')
    ws6.append(['ID', 'Musteri', 'Mulk', 'Tarih', 'Notlar', 'Durum'])
    for y in YerGosterme.query.filter_by(emlakci_id=emlakci.id).all():
        musteri_ad = Musteri.query.get(y.musteri_id).ad_soyad if y.musteri_id and Musteri.query.get(y.musteri_id) else ''
        mulk_ad = Mulk.query.get(y.mulk_id).baslik if y.mulk_id and Mulk.query.get(y.mulk_id) else ''
        ws6.append([y.id, musteri_ad, mulk_ad, str(y.tarih or y.olusturma or ''), y.notlar, y.durum])

    # Faturalar
    ws7 = wb.create_sheet('Faturalar')
    ws7.append(['ID', 'Fatura No', 'Tip', 'Alici', 'Tutar', 'KDV', 'Toplam', 'Durum', 'Tarih'])
    for f in Fatura.query.filter_by(emlakci_id=emlakci.id).all():
        ws7.append([f.id, f.fatura_no, f.tip, f.alici_ad, f.tutar, f.kdv_tutar, f.toplam, f.durum, str(f.olusturma or '')])

    # Cari Hesaplar
    ws8 = wb.create_sheet('Cariler')
    ws8.append(['ID', 'Ad', 'Tip', 'Bakiye', 'Tarih'])
    for c in Cari.query.filter_by(emlakci_id=emlakci.id).all():
        ws8.append([c.id, c.ad, c.tip, c.bakiye, str(c.olusturma or '')])

    # Leadler
    ws9 = wb.create_sheet('Leadler')
    ws9.append(['ID', 'Ad Soyad', 'Telefon', 'Kaynak', 'Durum', 'Puan', 'Notlar', 'Tarih'])
    for l in Lead.query.filter_by(emlakci_id=emlakci.id).all():
        ws9.append([l.id, l.ad_soyad, l.telefon, l.kaynak, l.durum, l.puan, l.notlar, str(l.olusturma or '')])

    # İletişim Geçmişi
    ws10 = wb.create_sheet('IletisimGecmisi')
    ws10.append(['ID', 'Musteri', 'Tip', 'Yon', 'Ozet', 'Tarih'])
    for k in IletisimKayit.query.filter_by(emlakci_id=emlakci.id).order_by(IletisimKayit.olusturma.desc()).limit(500).all():
        musteri_ad = Musteri.query.get(k.musteri_id).ad_soyad if k.musteri_id and Musteri.query.get(k.musteri_id) else ''
        ws10.append([k.id, musteri_ad, k.tip, k.yon, k.ozet, str(k.olusturma or '')])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def portfoy_excel_export(emlakci):
    """Sadece portföy verisini Excel olarak export et."""
    try:
        import openpyxl
    except ImportError:
        return json.dumps(_portfoy_json(emlakci), ensure_ascii=False).encode()

    from app.models import Musteri, Mulk
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Portfoy'
    ws.append(['ID', 'Baslik', 'Adres', 'Sehir', 'Ilce', 'Tip', 'Islem', 'Fiyat', 'Metrekare', 'Oda', 'Kat', 'Bina Yasi', 'Isitma', 'Esyali', 'Krediye Uygun', 'Sahibi', 'Sahip Tel', 'Tarih'])
    for m in Mulk.query.filter_by(emlakci_id=emlakci.id, aktif=True).all():
        d = m.detaylar or {}
        sahip_ad, sahip_tel = '', ''
        if m.musteri_id:
            sahip = Musteri.query.get(m.musteri_id)
            if sahip:
                sahip_ad, sahip_tel = sahip.ad_soyad or '', sahip.telefon or ''
        ws.append([m.id, m.baslik, m.adres, m.sehir, m.ilce, m.tip, m.islem_turu, m.fiyat, m.metrekare, m.oda_sayisi,
                   d.get('kat'), d.get('bina_yasi'), d.get('isitma'), d.get('esyali'), d.get('krediye_uygun'),
                   sahip_ad, sahip_tel, str(m.olusturma or '')])
    # Sütun genişlikleri
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def musteri_excel_export(emlakci):
    """Sadece müşteri verisini Excel olarak export et."""
    try:
        import openpyxl
    except ImportError:
        return json.dumps(_musteri_json(emlakci), ensure_ascii=False).encode()

    from app.models import Musteri
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Musteriler'
    ws.append(['ID', 'Ad Soyad', 'Telefon', 'Email', 'TC', 'Islem Turu', 'Butce Min', 'Butce Max', 'Sicaklik', 'Tercihler', 'Tarih'])
    for m in Musteri.query.filter_by(emlakci_id=emlakci.id).all():
        ws.append([m.id, m.ad_soyad, m.telefon, m.email, m.tc_kimlik, m.islem_turu, m.butce_min, m.butce_max, m.sicaklik, m.tercih_notlar, str(m.olusturma or '')])
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _portfoy_json(emlakci):
    from app.models import Mulk
    return [{'id': m.id, 'baslik': m.baslik, 'adres': m.adres, 'sehir': m.sehir, 'ilce': m.ilce,
             'tip': m.tip, 'islem': m.islem_turu, 'fiyat': m.fiyat, 'metrekare': m.metrekare,
             'oda': m.oda_sayisi} for m in Mulk.query.filter_by(emlakci_id=emlakci.id, aktif=True).all()]


def _musteri_json(emlakci):
    from app.models import Musteri
    return [{'id': m.id, 'ad_soyad': m.ad_soyad, 'telefon': m.telefon, 'islem': m.islem_turu,
             'butce_min': m.butce_min, 'butce_max': m.butce_max, 'sicaklik': m.sicaklik}
            for m in Musteri.query.filter_by(emlakci_id=emlakci.id).all()]


def _json_export(emlakci):
    """openpyxl yoksa JSON export."""
    from app.models import Musteri, Mulk, Not
    from app.models.muhasebe import GelirGider
    from app.models.planlama import Gorev

    data = {
        'emlakci': {'ad_soyad': emlakci.ad_soyad, 'email': emlakci.email, 'telefon': emlakci.telefon},
        'tarih': datetime.utcnow().isoformat(),
        'musteriler': [{'ad_soyad': m.ad_soyad, 'telefon': m.telefon, 'islem_turu': m.islem_turu} for m in Musteri.query.filter_by(emlakci_id=emlakci.id).all()],
        'mulkler': [{'baslik': m.baslik, 'adres': m.adres, 'fiyat': m.fiyat, 'tip': m.tip} for m in Mulk.query.filter_by(emlakci_id=emlakci.id, aktif=True).all()],
        'gelir_gider': [{'tip': k.tip, 'tutar': k.tutar, 'kategori': k.kategori} for k in GelirGider.query.filter_by(emlakci_id=emlakci.id).all()],
        'gorevler': [{'baslik': g.baslik, 'durum': g.durum} for g in Gorev.query.filter_by(emlakci_id=emlakci.id).all()],
    }
    return json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')


def yedek_durumu(emlakci):
    """Yedekleme durumu ve hatırlatma kontrolü."""
    from app.models import IslemLog
    son_yedek = IslemLog.query.filter_by(
        emlakci_id=emlakci.id, islem_tipi='yedekleme'
    ).order_by(IslemLog.olusturma.desc()).first()

    son_tarih = son_yedek.olusturma if son_yedek else None
    gun_gecti = (datetime.utcnow() - son_tarih).days if son_tarih else 999

    return {
        'son_yedek': son_tarih.isoformat() if son_tarih else None,
        'gun_gecti': gun_gecti,
        'uyari': gun_gecti >= 7,
        'kritik': gun_gecti >= 30,
        'mesaj': 'Yedek güncel' if gun_gecti < 7 else f'{gun_gecti} gündür yedek alınmadı!' if gun_gecti < 30 else 'KRİTİK: 30+ gündür yedek yok!',
    }


def yedek_logla(emlakci):
    """Yedek alındığında log kaydet."""
    from app.models import IslemLog
    log = IslemLog(emlakci_id=emlakci.id, islem_tipi='yedekleme', aciklama='Manuel yedek alındı')
    db.session.add(log)
    db.session.commit()


def depolama_durumu(emlakci):
    """Kullanıcının veri alanı kullanımı (tahmini)."""
    from app.models import Musteri, Mulk, Not
    from app.models.muhasebe import GelirGider
    from app.models.planlama import Gorev
    from app.models.islem_takip import Evrak

    sayilar = {
        'musteri': Musteri.query.filter_by(emlakci_id=emlakci.id).count(),
        'mulk': Mulk.query.filter_by(emlakci_id=emlakci.id).count(),
        'muhasebe': GelirGider.query.filter_by(emlakci_id=emlakci.id).count(),
        'gorev': Gorev.query.filter_by(emlakci_id=emlakci.id).count(),
        'evrak': Evrak.query.filter_by(emlakci_id=emlakci.id).count(),
        'not': Not.query.filter_by(emlakci_id=emlakci.id).count(),
    }

    # Tahmini boyut (kayıt başına ~1KB)
    toplam_kayit = sum(sayilar.values())
    tahmini_kb = toplam_kayit * 1
    tahmini_mb = round(tahmini_kb / 1024, 2)

    # Limit: 50MB (ücretsiz)
    limit_mb = 50
    doluluk_yuzde = round(tahmini_mb / limit_mb * 100, 1)

    return {
        'sayilar': sayilar,
        'toplam_kayit': toplam_kayit,
        'tahmini_mb': tahmini_mb,
        'limit_mb': limit_mb,
        'doluluk_yuzde': min(doluluk_yuzde, 100),
        'uyari': doluluk_yuzde > 80,
        'kritik': doluluk_yuzde > 95,
        'mesaj': 'Normal' if doluluk_yuzde < 80 else 'Yedek alın, alan dolmak üzere!' if doluluk_yuzde < 95 else 'KRİTİK: Alan dolu! Eski belgeleri silin.',
    }


def yedek_ozeti(emlakci):
    """Yedek için veri özeti."""
    from app.models import Musteri, Mulk, Not
    from app.models.muhasebe import GelirGider
    from app.models.planlama import Gorev

    return {
        'musteriler': Musteri.query.filter_by(emlakci_id=emlakci.id).count(),
        'mulkler': Mulk.query.filter_by(emlakci_id=emlakci.id, aktif=True).count(),
        'gelir_gider': GelirGider.query.filter_by(emlakci_id=emlakci.id).count(),
        'gorevler': Gorev.query.filter_by(emlakci_id=emlakci.id).count(),
        'notlar': Not.query.filter_by(emlakci_id=emlakci.id).count(),
    }
