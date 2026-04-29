"""
TOPLU İŞLEM SERVİSİ — Excel import, rehber import, OCR portföy
"""
import io
import json
import logging
from app.models import db, Musteri, Mulk
from app.services.ocr import fis_oku

logger = logging.getLogger(__name__)

# OCR prompt for sahibinden.com screenshots
_PORTFOY_OCR_PROMPT = """Bu bir emlak ilan listesi ekran görüntüsü (sahibinden.com veya benzeri).
Her ilan için aşağıdaki bilgileri JSON dizisi olarak çıkar:

[
  {
    "baslik": "ilan başlığı",
    "fiyat": 0,
    "adres": "konum/adres",
    "oda_sayisi": "3+1",
    "metrekare": 0,
    "islem_turu": "kira veya satis",
    "tip": "daire/villa/arsa/dukkan/ofis"
  }
]

Kurallar:
- Sadece JSON dizisi döndür
- Fiyatı sayı olarak yaz (TL işareti olmadan)
- islem_turu: "kira" veya "satis"
- Göremediğin alanları null yap
"""


def excel_musteri_import(emlakci_id, dosya_bytes):
    """Excel'den toplu müşteri import."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(dosya_bytes))
        ws = wb.active
    except Exception as e:
        return {'hata': f'Excel okunamadı: {e}', 'eklenen': 0}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    eklenen = 0

    for row in rows:
        if not row or not row[0]:
            continue
        ad_soyad = str(row[0]).strip()
        telefon = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        islem_turu = str(row[2]).strip().lower() if len(row) > 2 and row[2] else 'kira'
        if 'sat' in islem_turu:
            islem_turu = 'satis'
        else:
            islem_turu = 'kira'

        m = Musteri(emlakci_id=emlakci_id, ad_soyad=ad_soyad, telefon=telefon, islem_turu=islem_turu)
        db.session.add(m)
        eklenen += 1

    db.session.commit()
    return {'eklenen': eklenen, 'toplam_satir': len(rows)}


def excel_portfoy_import(emlakci_id, dosya_bytes):
    """Excel'den toplu portföy import."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(dosya_bytes))
        ws = wb.active
    except Exception as e:
        return {'hata': f'Excel okunamadı: {e}', 'eklenen': 0}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    eklenen = 0

    for row in rows:
        if not row or not row[0]:
            continue
        baslik = str(row[0]).strip()
        adres = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        tip = str(row[2]).strip().lower() if len(row) > 2 and row[2] else 'daire'
        islem = str(row[3]).strip().lower() if len(row) > 3 and row[3] else 'kira'
        fiyat = None
        if len(row) > 4 and row[4]:
            try: fiyat = float(str(row[4]).replace('.', '').replace(',', '.').replace('TL', '').strip())
            except: pass

        if 'sat' in islem: islem = 'satis'
        else: islem = 'kira'

        m = Mulk(emlakci_id=emlakci_id, baslik=baslik, adres=adres, tip=tip, islem_turu=islem, fiyat=fiyat)
        db.session.add(m)
        eklenen += 1

    db.session.commit()
    return {'eklenen': eklenen, 'toplam_satir': len(rows)}


def ocr_portfoy_import(emlakci_id, image_base64):
    """Sahibinden ekran görüntüsünden portföy import."""
    import os, requests

    gemini_key = os.environ.get('GEMINI_API_KEY', '')
    if not gemini_key:
        return {'hata': 'Gemini API anahtarı yok', 'eklenen': 0}

    try:
        url = f'https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={gemini_key}'
        payload = {
            'contents': [{'parts': [
                {'text': _PORTFOY_OCR_PROMPT},
                {'inline_data': {'mime_type': 'image/jpeg', 'data': image_base64}},
            ]}],
            'generationConfig': {'temperature': 0.1, 'maxOutputTokens': 2048},
        }
        r = requests.post(url, json=payload, timeout=30)
        r.raise_for_status()
        text = r.json()['candidates'][0]['content']['parts'][0]['text'].strip()
        if text.startswith('```'):
            text = text.split('\n', 1)[1].rsplit('```', 1)[0]
        ilanlar = json.loads(text)
    except Exception as e:
        return {'hata': f'OCR hatası: {e}', 'eklenen': 0}

    eklenen = 0
    for ilan in ilanlar:
        if not ilan.get('baslik'):
            continue
        m = Mulk(
            emlakci_id=emlakci_id,
            baslik=ilan.get('baslik', ''),
            adres=ilan.get('adres', ''),
            fiyat=ilan.get('fiyat'),
            oda_sayisi=ilan.get('oda_sayisi'),
            metrekare=ilan.get('metrekare'),
            islem_turu=ilan.get('islem_turu', 'satis'),
            tip=ilan.get('tip', 'daire'),
        )
        db.session.add(m)
        eklenen += 1

    db.session.commit()
    return {'eklenen': eklenen, 'toplam_ilan': len(ilanlar), 'ilanlar': ilanlar}


def rehber_import(emlakci_id, rehber_json):
    """Telefon rehberinden toplu müşteri ekleme. [{name, phone}]"""
    eklenen = 0
    for kisi in rehber_json:
        ad = kisi.get('name', '').strip()
        tel = kisi.get('phone', '').strip()
        if not ad:
            continue
        m = Musteri(emlakci_id=emlakci_id, ad_soyad=ad, telefon=tel)
        db.session.add(m)
        eklenen += 1

    db.session.commit()
    return {'eklenen': eklenen}
